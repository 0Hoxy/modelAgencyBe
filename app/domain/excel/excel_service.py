from io import BytesIO
from datetime import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from fastapi import HTTPException

from app.domain.models.models_repository import models_repository
from app.domain.models.models_schemas import ReadDomesticModel, ReadGlobalModel


class ExcelService:
    """엑셀 생성 서비스"""

    def _create_header_style(self):
        """헤더 스타일 생성"""
        return {
            'font': Font(bold=True, color="FFFFFF", size=12),
            'fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            'alignment': Alignment(horizontal="center", vertical="center"),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }

    def _create_cell_style(self):
        """일반 셀 스타일 생성"""
        return {
            'alignment': Alignment(horizontal="left", vertical="center"),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }

    def _apply_style(self, cell, style_dict):
        """셀에 스타일 적용"""
        for key, value in style_dict.items():
            setattr(cell, key, value)

    async def generate_domestic_excel(self) -> BytesIO:
        """
        국내 모델 엑셀 생성
        
        Returns:
            BytesIO: 엑셀 파일 바이트 스트림
            
        Raises:
            HTTPException: 데이터 조회 또는 엑셀 생성 실패
        """
        try:
            # 국내 모델 데이터 조회
            models_data = await models_repository.get_all_models_of_domestic()
            
            if not models_data:
                # 데이터가 없어도 빈 엑셀 생성
                models = []
            else:
                models = [ReadDomesticModel.model_validate(dict(record)) for record in models_data]

            # 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "국내 모델 목록"

            # 헤더 정의
            headers = [
                "ID", "이름", "예명", "생년월일", "성별", "전화번호", "국적",
                "소속사명", "소속사 담당자", "소속사 전화번호",
                "인스타그램", "틱톡", "유튜브",
                "도시", "구/군", "상세주소",
                "특기", "기타 언어", "타투 위치", "타투 크기",
                "키(cm)", "몸무게(kg)", "상의", "하의", "신발"
            ]

            # 헤더 스타일
            header_style = self._create_header_style()
            
            # 헤더 작성
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                self._apply_style(cell, header_style)

            # 데이터 작성
            cell_style = self._create_cell_style()
            
            for row_num, model in enumerate(models, 2):
                data = [
                    str(model.id),
                    model.name,
                    model.stage_name or "",
                    model.birth_date.strftime("%Y-%m-%d") if model.birth_date else "",
                    model.gender.to_korean() if model.gender else "",  # ✅ 한글 변환
                    model.phone or "",
                    model.nationality or "",
                    model.agency_name or "",
                    model.agency_manager_name or "",
                    model.agency_manager_phone or "",
                    model.instagram or "",
                    model.tictok or "",
                    model.youtube or "",
                    model.address_city or "",
                    model.address_district or "",
                    model.address_street or "",
                    model.special_abilities or "",
                    model.other_languages or "",
                    model.tattoo_location or "",
                    model.tattoo_size or "",
                    model.height if model.height else "",
                    model.weight if model.weight else "",
                    model.top_size or "",
                    model.bottom_size or "",
                    model.shoes_size or "",
                ]

                for col_num, value in enumerate(data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    self._apply_style(cell, cell_style)

            # 열 너비 자동 조정
            for col_num in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_num)
                ws.column_dimensions[column_letter].width = 15

            # 메모리에 저장
            excel_io = BytesIO()
            wb.save(excel_io)
            excel_io.seek(0)

            return excel_io

        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"국내 모델 엑셀 생성 중 오류가 발생했습니다: {str(e)}"
            )

    async def generate_global_excel(self) -> BytesIO:
        """
        해외 모델 엑셀 생성
        
        Returns:
            BytesIO: 엑셀 파일 바이트 스트림
            
        Raises:
            HTTPException: 데이터 조회 또는 엑셀 생성 실패
        """
        try:
            # 해외 모델 데이터 조회
            models_data = await models_repository.get_all_models_of_foreign()
            
            if not models_data:
                models = []
            else:
                models = [ReadGlobalModel.model_validate(dict(record)) for record in models_data]

            # 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "해외 모델 목록"

            # 헤더 정의
            headers = [
                "ID", "이름", "예명", "생년월일", "성별", "전화번호", "국적",
                "인스타그램", "유튜브", "카카오톡",
                "도시", "구/군", "상세주소",
                "특기", "모국어", "기타 언어", "한국어 수준",
                "타투 위치", "타투 크기", "비자 타입",
                "키(cm)", "몸무게(kg)", "상의", "하의", "신발"
            ]

            # 헤더 스타일
            header_style = self._create_header_style()
            
            # 헤더 작성
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                self._apply_style(cell, header_style)

            # 데이터 작성
            cell_style = self._create_cell_style()
            
            for row_num, model in enumerate(models, 2):
                data = [
                    str(model.id),
                    model.name,
                    model.stage_name or "",
                    model.birth_date.strftime("%Y-%m-%d") if model.birth_date else "",
                    model.gender.to_korean() if model.gender else "",  # ✅ 한글 변환
                    model.phone or "",
                    model.nationality or "",
                    model.instagram or "",
                    model.youtube or "",
                    model.kakaotalk or "",
                    model.address_city or "",
                    model.address_district or "",
                    model.address_street or "",
                    model.special_abilities or "",
                    model.first_language or "",
                    model.other_languages or "",
                    model.korean_level.to_korean() if model.korean_level else "",  # ✅ 한글 변환
                    model.tattoo_location or "",
                    model.tattoo_size or "",
                    model.visa_type.to_korean() if model.visa_type else "",  # ✅ 한글 변환
                    model.height if model.height else "",
                    model.weight if model.weight else "",
                    model.top_size or "",
                    model.bottom_size or "",
                    model.shoes_size or "",
                ]

                for col_num, value in enumerate(data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    self._apply_style(cell, cell_style)

            # 열 너비 자동 조정
            for col_num in range(1, len(headers) + 1):
                column_letter = get_column_letter(col_num)
                ws.column_dimensions[column_letter].width = 15

            # 메모리에 저장
            excel_io = BytesIO()
            wb.save(excel_io)
            excel_io.seek(0)

            return excel_io

        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"해외 모델 엑셀 생성 중 오류가 발생했습니다: {str(e)}"
            )


excel_service = ExcelService()
